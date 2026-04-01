#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import subprocess
import sys
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent


def run_command(args: list[str]) -> None:
    result = subprocess.run(args, cwd=ROOT, check=False)
    if result.returncode != 0:
        raise SystemExit(result.returncode)


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def clean_sheet_title(title: str) -> str:
    invalid = set(r'[]:*?/\\')
    cleaned = "".join("_" if char in invalid else char for char in title)
    return cleaned[:31]


def autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 38)


def style_table_sheet(ws) -> None:
    if ws.max_row == 0 or ws.max_column == 0:
        return
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_worksheet(ws)


def write_records_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet(clean_sheet_title(title))
    if not rows:
        ws["A1"] = "No rows"
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    style_table_sheet(ws)


def write_summary_sheet(
    workbook: Workbook,
    prediction_rows: list[dict[str, str]],
    pitcher_prop_rows: list[dict[str, str]],
    batter_prop_rows: list[dict[str, str]],
    run_date: date,
    workbook_path: Path,
) -> None:
    ws = workbook.active
    ws.title = "Summary"
    ws["A1"] = "MLB Export Workbook"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Generated"
    ws["B3"] = datetime.now(UTC).isoformat()
    ws["A4"] = "Slate Date"
    ws["B4"] = run_date.isoformat()
    ws["A5"] = "Workbook"
    ws["B5"] = str(workbook_path)

    ws["A7"] = "Game Predictions"
    ws["A7"].font = Font(size=13, bold=True)
    headers = [
        "Away Team",
        "Home Team",
        "Away Win %",
        "Home Win %",
        "Predicted Total",
        "Run Line Side",
        "Pitchers",
    ]
    for index, header in enumerate(headers, start=1):
        ws.cell(row=8, column=index, value=header)
    for cell in ws[8]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for row_index, row in enumerate(prediction_rows, start=9):
        ws.cell(row=row_index, column=1, value=row.get("away_team_name"))
        ws.cell(row=row_index, column=2, value=row.get("home_team_name"))
        ws.cell(row=row_index, column=3, value=row.get("away_win_probability"))
        ws.cell(row=row_index, column=4, value=row.get("home_win_probability"))
        ws.cell(row=row_index, column=5, value=row.get("predicted_total_runs"))
        ws.cell(row=row_index, column=6, value=row.get("predicted_runline_side"))
        ws.cell(
            row=row_index,
            column=7,
            value=f"{row.get('away_probable_pitcher_name', '')} vs {row.get('home_probable_pitcher_name', '')}",
        )

    start_row = 10 + max(len(prediction_rows), 1)
    ws.cell(row=start_row, column=1, value="Top Pitcher Props").font = Font(size=13, bold=True)
    prop_headers = ["Team", "Pitcher", "Pred Ks", "Pred Outs", "Pred ER", "Pred Hits Allowed"]
    for index, header in enumerate(prop_headers, start=1):
        ws.cell(row=start_row + 1, column=index, value=header)
    for cell in ws[start_row + 1]:
        cell.fill = PatternFill("solid", fgColor="375623")
        cell.font = Font(color="FFFFFF", bold=True)
    display_pitchers = pitcher_prop_rows[: min(12, len(pitcher_prop_rows))]
    for row_index, row in enumerate(display_pitchers, start=start_row + 2):
        away_pitcher = row.get("away_pitcher_name", "")
        home_pitcher = row.get("home_pitcher_name", "")
        ws.cell(row=row_index, column=1, value=row.get("away_team_name"))
        ws.cell(row=row_index, column=2, value=away_pitcher)
        ws.cell(row=row_index, column=3, value=row.get("away_pitcher_predicted_strikeouts"))
        ws.cell(row=row_index, column=4, value=row.get("away_pitcher_predicted_outs"))
        ws.cell(row=row_index, column=5, value=row.get("away_pitcher_predicted_earned_runs"))
        ws.cell(row=row_index, column=6, value=row.get("away_pitcher_predicted_hits_allowed"))
        ws.cell(row=row_index + len(display_pitchers), column=1, value=row.get("home_team_name"))
        ws.cell(row=row_index + len(display_pitchers), column=2, value=home_pitcher)
        ws.cell(row=row_index + len(display_pitchers), column=3, value=row.get("home_pitcher_predicted_strikeouts"))
        ws.cell(row=row_index + len(display_pitchers), column=4, value=row.get("home_pitcher_predicted_outs"))
        ws.cell(row=row_index + len(display_pitchers), column=5, value=row.get("home_pitcher_predicted_earned_runs"))
        ws.cell(row=row_index + len(display_pitchers), column=6, value=row.get("home_pitcher_predicted_hits_allowed"))

    batter_start = start_row + 4 + (2 * max(len(display_pitchers), 1))
    ws.cell(row=batter_start, column=1, value="Top Batter Props").font = Font(size=13, bold=True)
    batter_headers = ["Team", "Player", "1+ Hit %", "Exp Hits", "Exp TB", "Exp RBI", "HR %"]
    for index, header in enumerate(batter_headers, start=1):
        ws.cell(row=batter_start + 1, column=index, value=header)
    for cell in ws[batter_start + 1]:
        cell.fill = PatternFill("solid", fgColor="7F6000")
        cell.font = Font(color="FFFFFF", bold=True)
    top_batters = sorted(
        batter_prop_rows,
        key=lambda row: float(row.get("prop_score") or 0.0),
        reverse=True,
    )[:20]
    for row_index, row in enumerate(top_batters, start=batter_start + 2):
        ws.cell(row=row_index, column=1, value=row.get("team_name"))
        ws.cell(row=row_index, column=2, value=row.get("player_name"))
        ws.cell(row=row_index, column=3, value=row.get("hit_probability"))
        ws.cell(row=row_index, column=4, value=row.get("expected_hits"))
        ws.cell(row=row_index, column=5, value=row.get("expected_total_bases"))
        ws.cell(row=row_index, column=6, value=row.get("expected_rbi"))
        ws.cell(row=row_index, column=7, value=row.get("home_run_probability"))

    autosize_worksheet(ws)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run the MLB pipeline and export everything into one Excel workbook.")
    today = date.today()
    parser.add_argument("--date", default=today.isoformat(), help="Slate date in YYYY-MM-DD.")
    parser.add_argument("--history-start-date", default="2024-03-28", help="Historical feature window start date.")
    parser.add_argument("--training-start-date", default="2025-03-27", help="Start date for the training export window.")
    parser.add_argument("--season", type=int, default=today.year, help="Season year.")
    parser.add_argument("--recent-games", type=int, default=15)
    parser.add_argument("--pitcher-recent-games", type=int, default=6)
    parser.add_argument("--bullpen-recent-games", type=int, default=5)
    parser.add_argument("--market-weight", type=float, default=0.5)
    parser.add_argument("--output-dir", default=None, help="Directory for all CSVs and the Excel workbook.")
    return parser


def main() -> int:
    args = build_arg_parser().parse_args()
    run_date = date.fromisoformat(args.date)
    output_dir = Path(args.output_dir) if args.output_dir else ROOT / f"exports_{run_date.strftime('%Y%m%d')}"
    output_dir.mkdir(parents=True, exist_ok=True)

    training_dir = output_dir / "training"
    slate_dir = output_dir / "slate"
    prediction_dir = output_dir / "predictions"
    props_dir = output_dir / "props"
    for path in (training_dir, slate_dir, prediction_dir, props_dir):
        path.mkdir(parents=True, exist_ok=True)

    python_exe = sys.executable
    run_command(
        [
            python_exe,
            ".\\mlb_model_builder.py",
            "--start-date",
            args.training_start_date,
            "--end-date",
            args.date,
            "--season",
            str(args.season),
            "--history-start-date",
            args.history_start_date,
            "--recent-games",
            str(args.recent_games),
            "--pitcher-recent-games",
            str(args.pitcher_recent_games),
            "--bullpen-recent-games",
            str(args.bullpen_recent_games),
            "--include-weather",
            "--output-dir",
            str(training_dir),
        ]
    )
    run_command(
        [
            python_exe,
            ".\\mlb_model_builder.py",
            "--start-date",
            args.date,
            "--end-date",
            args.date,
            "--season",
            str(args.season),
            "--history-start-date",
            args.history_start_date,
            "--recent-games",
            str(args.recent_games),
            "--pitcher-recent-games",
            str(args.pitcher_recent_games),
            "--bullpen-recent-games",
            str(args.bullpen_recent_games),
            "--include-weather",
            "--output-dir",
            str(slate_dir),
        ]
    )
    run_command(
        [
            python_exe,
            ".\\mlb_predictor.py",
            "--input-csv",
            str(training_dir / "mlb_game_features.csv"),
            "--predict-csv",
            str(slate_dir / "mlb_game_features.csv"),
            "--output-dir",
            str(prediction_dir),
            "--prediction-start-date",
            args.date,
            "--fetch-cbs-odds",
            "--market-weight",
            str(args.market_weight),
        ]
    )
    run_command(
        [
            python_exe,
            ".\\mlb_player_props.py",
            "--predict-csv",
            str(slate_dir / "mlb_game_features.csv"),
            "--output-dir",
            str(props_dir),
        ]
    )

    prediction_rows = read_csv(prediction_dir / "mlb_predictions.csv")
    pitcher_prop_rows = read_csv(props_dir / "mlb_player_props.csv")
    batter_prop_rows = read_csv(props_dir / "mlb_batter_props.csv")
    game_feature_rows = read_csv(slate_dir / "mlb_game_features.csv")
    snapshot_rows = read_csv(slate_dir / "mlb_team_pitcher_snapshots.csv")

    workbook_path = output_dir / f"mlb_export_{run_date.strftime('%Y%m%d')}.xlsx"
    workbook = Workbook()
    write_summary_sheet(workbook, prediction_rows, pitcher_prop_rows, batter_prop_rows, run_date, workbook_path)
    write_records_sheet(workbook, "Game Predictions", prediction_rows)
    write_records_sheet(workbook, "Pitcher Props", pitcher_prop_rows)
    write_records_sheet(workbook, "Batter Props", batter_prop_rows)
    write_records_sheet(workbook, "Game Features", game_feature_rows)
    write_records_sheet(workbook, "Team Snapshots", snapshot_rows)
    workbook.save(workbook_path)

    print(f"Excel workbook written to {workbook_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
